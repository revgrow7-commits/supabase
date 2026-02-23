"""
Reports routes - Migrated from server.py
Handles all reporting endpoints for productivity, families, installers, etc.
"""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone
from io import BytesIO
import logging
import re

from database import db
from security import get_current_user, require_role
from models.user import User, UserRole

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter()
logger = logging.getLogger(__name__)


# ============ HELPER FUNCTIONS ============

def classify_product_to_family(product_name: str) -> tuple:
    """
    Classifies a product into a family based on keywords.
    Returns (family_name, confidence_percent).
    """
    if not product_name:
        return "Outros", 0
    
    name = product_name.lower()
    
    classifications = [
        (["adesivo", "vinil", "plotagem", "recorte", "adesivos"], "Adesivos", 90),
        (["lona", "banner", "faixa", "frontlight", "backlight"], "Lonas/Banners", 90),
        (["chapa", "acm", "fachada", "alumínio composto"], "Chapas/Fachadas", 85),
        (["placa", "legenda", "sinalização"], "Placas/Legendas", 80),
        (["display", "expositor", "totem", "stand"], "Displays/Totens", 85),
        (["serviço", "instalação", "entrega", "mão de obra"], "Serviços", 70),
        (["impressão", "impresso", "papel"], "Impressos", 75),
    ]
    
    for keywords, family, confidence in classifications:
        for keyword in keywords:
            if keyword in name:
                return family, confidence
    
    return "Outros", 30


def calculate_job_products_area(holdprint_data: dict) -> tuple:
    """
    Calculates total area for all products in a job.
    Returns (products_with_area, total_area_m2, total_products, total_quantity).
    """
    products_with_area = []
    total_area_m2 = 0
    total_products = 0
    total_quantity = 0
    
    products = holdprint_data.get("products", [])
    
    for product in products:
        product_name = product.get("name", "")
        quantity = product.get("quantity", 1)
        description = product.get("description", "")
        
        width_m = None
        height_m = None
        
        width_match = re.search(r'Largura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
        height_match = re.search(r'Altura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
        
        if width_match:
            width_m = float(width_match.group(1).replace(',', '.'))
        if height_match:
            height_m = float(height_match.group(1).replace(',', '.'))
        
        area_m2 = None
        if width_m and height_m:
            area_m2 = round(width_m * height_m * quantity, 2)
            total_area_m2 += area_m2
        
        family_name, confidence = classify_product_to_family(product_name)
        
        products_with_area.append({
            "name": product_name,
            "quantity": quantity,
            "width_m": width_m,
            "height_m": height_m,
            "total_area_m2": area_m2,
            "family_name": family_name,
            "family_confidence": confidence,
            "unit_price": product.get("unitPrice", 0),
            "total_value": product.get("totalValue", 0)
        })
        
        total_products += 1
        total_quantity += quantity
    
    return products_with_area, round(total_area_m2, 2), total_products, total_quantity


# ============ ROUTES ============

@router.get("/reports/by-family")
async def get_report_by_family(current_user: User = Depends(get_current_user)):
    """
    Relatório completo por família de produtos.
    Analisa todos os jobs importados e classifica seus produtos por família.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    families = await db.product_families.find({}, {"_id": 0}).to_list(100)
    family_map = {f["name"]: f for f in families}
    
    family_report = {}
    all_products = []
    unclassified_products = []
    
    for job in jobs:
        holdprint_data = job.get("holdprint_data", {})
        products = holdprint_data.get("products", [])
        production_items = holdprint_data.get("production", {}).get("items", [])
        
        for product in products:
            product_name = product.get("name", "")
            quantity = product.get("quantity", 1)
            description = product.get("description", "")
            
            width_m = None
            height_m = None
            
            width_match = re.search(r'Largura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
            height_match = re.search(r'Altura:\s*<span[^>]*>([0-9.,]+)\s*m', description, re.IGNORECASE)
            
            if width_match:
                width_m = float(width_match.group(1).replace(',', '.'))
            if height_match:
                height_m = float(height_match.group(1).replace(',', '.'))
            
            area_m2 = None
            if width_m and height_m:
                area_m2 = round(width_m * height_m * quantity, 2)
            
            family_name, confidence = classify_product_to_family(product_name)
            
            product_data = {
                "job_id": job.get("id"),
                "job_title": job.get("title"),
                "job_code": holdprint_data.get("code"),
                "client_name": holdprint_data.get("customerName", job.get("client_name")),
                "product_name": product_name,
                "family_name": family_name,
                "confidence": confidence,
                "quantity": quantity,
                "width_m": width_m,
                "height_m": height_m,
                "area_m2": area_m2,
                "unit_price": product.get("unitPrice", 0),
                "total_value": product.get("totalValue", 0),
                "branch": job.get("branch")
            }
            
            all_products.append(product_data)
            
            if family_name not in family_report:
                family_info = family_map.get(family_name, {})
                family_report[family_name] = {
                    "family_name": family_name,
                    "color": family_info.get("color", "#6B7280"),
                    "total_jobs": set(),
                    "total_products": 0,
                    "total_quantity": 0,
                    "total_area_m2": 0,
                    "total_value": 0,
                    "products": []
                }
            
            family_report[family_name]["total_jobs"].add(job.get("id"))
            family_report[family_name]["total_products"] += 1
            family_report[family_name]["total_quantity"] += quantity
            if area_m2:
                family_report[family_name]["total_area_m2"] += area_m2
            family_report[family_name]["total_value"] += product.get("totalValue", 0)
            family_report[family_name]["products"].append(product_data)
            
            if confidence < 50:
                unclassified_products.append(product_data)
        
        for item in production_items:
            item_name = item.get("name", "")
            item_quantity = item.get("quantity", 1)
            
            family_name, confidence = classify_product_to_family(item_name)
            
            if family_name not in family_report:
                family_info = family_map.get(family_name, {})
                family_report[family_name] = {
                    "family_name": family_name,
                    "color": family_info.get("color", "#6B7280"),
                    "total_jobs": set(),
                    "total_products": 0,
                    "total_quantity": 0,
                    "total_area_m2": 0,
                    "total_value": 0,
                    "products": []
                }
            
            family_report[family_name]["total_jobs"].add(job.get("id"))
            family_report[family_name]["total_quantity"] += item_quantity
    
    for family_name in family_report:
        family_report[family_name]["total_jobs"] = len(family_report[family_name]["total_jobs"])
        family_report[family_name]["total_area_m2"] = round(family_report[family_name]["total_area_m2"], 2)
        family_report[family_name]["total_value"] = round(family_report[family_name]["total_value"], 2)
        family_report[family_name]["products"] = family_report[family_name]["products"][:50]
    
    sorted_families = sorted(family_report.values(), key=lambda x: x["total_quantity"], reverse=True)
    
    total_area = sum(f["total_area_m2"] for f in sorted_families)
    total_value = sum(f["total_value"] for f in sorted_families)
    total_products = sum(f["total_products"] for f in sorted_families)
    
    return {
        "summary": {
            "total_jobs": len(jobs),
            "total_products": total_products,
            "total_area_m2": round(total_area, 2),
            "total_value": round(total_value, 2),
            "families_count": len(sorted_families),
            "unclassified_count": len(unclassified_products)
        },
        "by_family": sorted_families,
        "unclassified": unclassified_products[:20],
        "all_products": all_products[:100]
    }


@router.get("/reports/kpis/family-productivity")
async def get_family_productivity_kpis(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user)
):
    """
    KPIs de produtividade por família de produto.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    query = {"status": "completed"}
    if date_from or date_to:
        date_filter = {}
        if date_from:
            date_filter["$gte"] = date_from + "T00:00:00"
        if date_to:
            date_filter["$lte"] = date_to + "T23:59:59"
        if date_filter:
            query["checkin_at"] = date_filter
    
    checkins = await db.item_checkins.find(query, {"_id": 0}).to_list(10000)
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    jobs_map = {j["id"]: j for j in jobs}
    
    family_data = {}
    global_totals = {"total_m2": 0, "total_minutes": 0, "count": 0}
    
    for checkin in checkins:
        family = checkin.get("family_name") or "Outros"
        installed_m2 = checkin.get("installed_m2", 0) or 0
        duration = checkin.get("net_duration_minutes") or checkin.get("duration_minutes", 0) or 0
        
        if installed_m2 <= 0:
            continue
            
        if family not in family_data:
            family_data[family] = {
                "family_name": family,
                "total_m2": 0,
                "total_minutes": 0,
                "count": 0,
                "min_duration": float('inf'),
                "max_duration": 0,
                "m2_list": [],
                "duration_list": [],
                "installers": set(),
                "jobs": set()
            }
        
        family_data[family]["total_m2"] += installed_m2
        family_data[family]["total_minutes"] += duration
        family_data[family]["count"] += 1
        family_data[family]["m2_list"].append(installed_m2)
        family_data[family]["duration_list"].append(duration)
        
        if duration > 0:
            family_data[family]["min_duration"] = min(family_data[family]["min_duration"], duration)
            family_data[family]["max_duration"] = max(family_data[family]["max_duration"], duration)
        
        if checkin.get("installer_id"):
            family_data[family]["installers"].add(checkin.get("installer_id"))
        if checkin.get("job_id"):
            family_data[family]["jobs"].add(checkin.get("job_id"))
        
        global_totals["total_m2"] += installed_m2
        global_totals["total_minutes"] += duration
        global_totals["count"] += 1
    
    result = []
    global_avg_m2_h = (global_totals["total_m2"] / global_totals["total_minutes"] * 60) if global_totals["total_minutes"] > 0 else 0
    
    for family, data in family_data.items():
        avg_m2_per_hour = (data["total_m2"] / data["total_minutes"] * 60) if data["total_minutes"] > 0 else 0
        avg_m2_per_install = data["total_m2"] / data["count"] if data["count"] > 0 else 0
        avg_duration = data["total_minutes"] / data["count"] if data["count"] > 0 else 0
        efficiency = (avg_m2_per_hour / global_avg_m2_h * 100) if global_avg_m2_h > 0 else 100
        
        if len(data["m2_list"]) > 1:
            mean_m2 = sum(data["m2_list"]) / len(data["m2_list"])
            variance = sum((x - mean_m2) ** 2 for x in data["m2_list"]) / len(data["m2_list"])
            std_dev_m2 = variance ** 0.5
        else:
            std_dev_m2 = 0
        
        result.append({
            "family_name": family,
            "total_m2": round(data["total_m2"], 2),
            "total_hours": round(data["total_minutes"] / 60, 2),
            "installation_count": data["count"],
            "avg_m2_per_hour": round(avg_m2_per_hour, 2),
            "avg_m2_per_install": round(avg_m2_per_install, 2),
            "avg_duration_minutes": round(avg_duration, 1),
            "min_duration_minutes": data["min_duration"] if data["min_duration"] != float('inf') else 0,
            "max_duration_minutes": data["max_duration"],
            "efficiency_percent": round(efficiency, 1),
            "std_dev_m2": round(std_dev_m2, 2),
            "unique_installers": len(data["installers"]),
            "unique_jobs": len(data["jobs"]),
            "share_of_total_m2": round(data["total_m2"] / global_totals["total_m2"] * 100, 1) if global_totals["total_m2"] > 0 else 0
        })
    
    result.sort(key=lambda x: x["total_m2"], reverse=True)
    
    for i, item in enumerate(result, 1):
        item["rank"] = i
    
    return {
        "kpis": result,
        "summary": {
            "total_families": len(result),
            "global_total_m2": round(global_totals["total_m2"], 2),
            "global_total_hours": round(global_totals["total_minutes"] / 60, 2),
            "global_installations": global_totals["count"],
            "global_avg_m2_per_hour": round(global_avg_m2_h, 2),
            "period": {"from": date_from, "to": date_to}
        }
    }


@router.get("/reports/by-installer")
async def get_report_by_installer(current_user: User = Depends(get_current_user)):
    """
    Relatório de produtividade por instalador.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    item_checkins = await db.item_checkins.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    
    jobs_map = {job["id"]: job for job in jobs}
    installer_report = []
    
    for installer in installers:
        installer_id = installer["id"]
        installer_checkins = [c for c in item_checkins if c.get("installer_id") == installer_id]
        
        completed_count = len(installer_checkins)
        total_net_duration_min = 0
        total_m2_installed = 0
        
        for checkin in installer_checkins:
            net_minutes = checkin.get("net_duration_minutes") or checkin.get("duration_minutes") or 0
            total_net_duration_min += net_minutes
            
            job = jobs_map.get(checkin.get("job_id"))
            if job:
                products = job.get("products_with_area", [])
                item_index = checkin.get("item_index", 0)
                if item_index < len(products):
                    item = products[item_index]
                    item_m2 = item.get("total_area_m2", 0) or 0
                    total_m2_installed += item_m2
        
        job_ids = set(c.get("job_id") for c in installer_checkins if c.get("job_id"))
        jobs_worked = len(job_ids)
        
        jobs_details = []
        for job_id in job_ids:
            job = jobs_map.get(job_id)
            if job:
                job_area = job.get("area_m2", 0) or 0
                job_item_checkins = [c for c in installer_checkins if c.get("job_id") == job_id]
                job_net_duration = sum(c.get("net_duration_minutes") or c.get("duration_minutes") or 0 for c in job_item_checkins)
                
                job_m2_installed = 0
                products = job.get("products_with_area", [])
                for checkin in job_item_checkins:
                    item_index = checkin.get("item_index", 0)
                    if item_index < len(products):
                        job_m2_installed += products[item_index].get("total_area_m2", 0) or 0
                
                jobs_details.append({
                    "job_id": job_id,
                    "job_title": job.get("title"),
                    "client": job.get("client_name") or job.get("holdprint_data", {}).get("customerName"),
                    "job_area_m2": job_area,
                    "duration_min": round(job_net_duration, 2),
                    "m2_installed": round(job_m2_installed, 2),
                    "status": job.get("status"),
                    "items_completed": len(job_item_checkins)
                })
        
        productivity_m2_h = 0
        total_hours = total_net_duration_min / 60 if total_net_duration_min > 0 else 0
        if total_hours > 0 and total_m2_installed > 0:
            productivity_m2_h = round(total_m2_installed / total_hours, 2)
        
        installer_data = {
            "installer_id": installer_id,
            "full_name": installer.get("full_name"),
            "branch": installer.get("branch"),
            "metrics": {
                "items_completed": completed_count,
                "completed_checkins": completed_count,
                "jobs_worked": jobs_worked,
                "total_duration_hours": round(total_hours, 2),
                "total_m2_reported": round(total_m2_installed, 2),
                "productivity_m2_h": productivity_m2_h
            },
            "jobs": sorted(jobs_details, key=lambda x: x.get("m2_installed", 0), reverse=True)[:20]
        }
        
        installer_report.append(installer_data)
    
    installer_report.sort(key=lambda x: x["metrics"]["productivity_m2_h"], reverse=True)
    
    total_area_all = sum(i["metrics"]["total_m2_reported"] for i in installer_report)
    total_hours_all = sum(i["metrics"]["total_duration_hours"] for i in installer_report)
    
    return {
        "summary": {
            "total_installers": len(installer_report),
            "total_area_m2_all": round(total_area_all, 2),
            "total_hours_all": round(total_hours_all, 2),
            "avg_productivity_m2_h": round(total_area_all / total_hours_all, 2) if total_hours_all > 0 else 0
        },
        "by_installer": installer_report
    }


@router.get("/reports/productivity")
async def get_productivity_report(
    filter_by: Optional[str] = Query(None, description="Filter type: installer, job, family, item"),
    filter_id: Optional[str] = Query(None, description="ID to filter by"),
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    current_user: User = Depends(get_current_user)
):
    """
    Relatório de produtividade completo.
    """
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(10000)
    item_checkins = await db.item_checkins.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    legacy_checkins = await db.checkins.find({"status": "completed"}, {"_id": 0}).to_list(10000)
    
    jobs_map = {job["id"]: job for job in jobs}
    installers_map = {inst["id"]: inst for inst in installers}
    
    by_installer = {}
    by_job = {}
    by_family = {}
    by_item = {}
    
    for checkin in item_checkins:
        job = jobs_map.get(checkin.get("job_id"))
        if not job:
            continue
        
        checkin_at = checkin.get("checkin_at")
        if isinstance(checkin_at, str):
            checkin_at = datetime.fromisoformat(checkin_at.replace('Z', '+00:00'))
        
        if date_from:
            start_date = datetime.fromisoformat(date_from + "T00:00:00+00:00")
            if checkin_at and checkin_at < start_date:
                continue
        
        if date_to:
            end_date = datetime.fromisoformat(date_to + "T23:59:59+00:00")
            if checkin_at and checkin_at > end_date:
                continue
        
        products = job.get("products_with_area", [])
        item_index = checkin.get("item_index", 0)
        item = products[item_index] if item_index < len(products) else {}
        
        item_m2 = item.get("total_area_m2", 0) or 0
        
        checkout_at = checkin.get("checkout_at")
        if isinstance(checkout_at, str):
            checkout_at = datetime.fromisoformat(checkout_at.replace('Z', '+00:00'))
        
        if checkin_at and checkin_at.tzinfo is None:
            checkin_at = checkin_at.replace(tzinfo=timezone.utc)
        if checkout_at and checkout_at.tzinfo is None:
            checkout_at = checkout_at.replace(tzinfo=timezone.utc)
        
        net_duration_minutes = checkin.get("net_duration_minutes")
        total_pause_minutes = checkin.get("total_pause_minutes", 0) or 0
        
        if net_duration_minutes is None:
            if checkin_at and checkout_at:
                net_duration_minutes = (checkout_at - checkin_at).total_seconds() / 60
            else:
                net_duration_minutes = 0
        
        duration_minutes = net_duration_minutes
        
        installer_id = checkin.get("installer_id")
        installer = installers_map.get(installer_id, {})
        installer_name = installer.get("full_name", "Desconhecido")
        family_name = item.get("family_name", "Não Classificado")
        
        record = {
            "job_id": job.get("id"),
            "job_title": job.get("title"),
            "client_name": job.get("client_name") or job.get("holdprint_data", {}).get("customerName"),
            "installer_id": installer_id,
            "installer_name": installer_name,
            "item_name": item.get("name", f"Item {item_index + 1}"),
            "item_index": item_index,
            "family_name": family_name,
            "m2_api": item_m2,
            "m2_reported": checkin.get("installed_m2", 0) or 0,
            "duration_minutes": round(duration_minutes, 2),
            "gross_duration_minutes": checkin.get("duration_minutes", 0) or 0,
            "pause_minutes": total_pause_minutes,
            "checkin_at": checkin_at.isoformat() if checkin_at else None,
            "checkout_at": checkout_at.isoformat() if checkout_at else None,
            "complexity_level": checkin.get("complexity_level"),
            "scenario_category": checkin.get("scenario_category"),
            "notes": checkin.get("notes")
        }
        
        if filter_by == "installer" and filter_id and installer_id != filter_id:
            continue
        if filter_by == "job" and filter_id and job.get("id") != filter_id:
            continue
        if filter_by == "family" and filter_id and family_name != filter_id:
            continue
        
        # Aggregate by installer
        if installer_id not in by_installer:
            by_installer[installer_id] = {
                "installer_id": installer_id,
                "installer_name": installer_name,
                "branch": installer.get("branch"),
                "total_m2": 0,
                "total_minutes": 0,
                "items_count": 0,
                "jobs": set(),
                "records": []
            }
        by_installer[installer_id]["total_m2"] += item_m2
        by_installer[installer_id]["total_minutes"] += duration_minutes
        by_installer[installer_id]["items_count"] += 1
        by_installer[installer_id]["jobs"].add(job.get("id"))
        by_installer[installer_id]["records"].append(record)
        
        # Aggregate by job
        job_id = job.get("id")
        if job_id not in by_job:
            by_job[job_id] = {
                "job_id": job_id,
                "job_title": job.get("title"),
                "client_name": job.get("client_name") or job.get("holdprint_data", {}).get("customerName"),
                "total_m2_api": job.get("area_m2", 0) or 0,
                "total_m2_executed": 0,
                "total_minutes": 0,
                "items_count": 0,
                "installers": set(),
                "records": []
            }
        by_job[job_id]["total_m2_executed"] += item_m2
        by_job[job_id]["total_minutes"] += duration_minutes
        by_job[job_id]["items_count"] += 1
        by_job[job_id]["installers"].add(installer_id)
        by_job[job_id]["records"].append(record)
        
        # Aggregate by family
        if family_name not in by_family:
            by_family[family_name] = {
                "family_name": family_name,
                "total_m2": 0,
                "total_minutes": 0,
                "items_count": 0,
                "jobs": set(),
                "installers": set(),
                "records": []
            }
        by_family[family_name]["total_m2"] += item_m2
        by_family[family_name]["total_minutes"] += duration_minutes
        by_family[family_name]["items_count"] += 1
        by_family[family_name]["jobs"].add(job.get("id"))
        by_family[family_name]["installers"].add(installer_id)
        by_family[family_name]["records"].append(record)
        
        # Aggregate by item
        item_key = f"{job_id}:{item_index}"
        if item_key not in by_item:
            by_item[item_key] = {
                "job_id": job_id,
                "job_title": job.get("title"),
                "item_index": item_index,
                "item_name": item.get("name", f"Item {item_index + 1}"),
                "family_name": family_name,
                "m2_api": item_m2,
                "total_minutes": 0,
                "executions": 0,
                "installers": set(),
                "records": []
            }
        by_item[item_key]["total_minutes"] += duration_minutes
        by_item[item_key]["executions"] += 1
        by_item[item_key]["installers"].add(installer_id)
        by_item[item_key]["records"].append(record)
    
    # Process legacy checkins
    for checkin in legacy_checkins:
        job = jobs_map.get(checkin.get("job_id"))
        if not job:
            continue
        
        checkin_at = checkin.get("checkin_at")
        if isinstance(checkin_at, str):
            checkin_at = datetime.fromisoformat(checkin_at.replace('Z', '+00:00'))
        
        if date_from:
            start_date = datetime.fromisoformat(date_from + "T00:00:00+00:00")
            if checkin_at and checkin_at < start_date:
                continue
        
        if date_to:
            end_date = datetime.fromisoformat(date_to + "T23:59:59+00:00")
            if checkin_at and checkin_at > end_date:
                continue
        
        duration_minutes = checkin.get("duration_minutes", 0) or 0
        installer_id = checkin.get("installer_id")
        installer = installers_map.get(installer_id, {})
        installer_name = installer.get("full_name", "Desconhecido")
        
        if filter_by == "installer" and filter_id and installer_id != filter_id:
            continue
        if filter_by == "job" and filter_id and job.get("id") != filter_id:
            continue
        
        if installer_id not in by_installer:
            by_installer[installer_id] = {
                "installer_id": installer_id,
                "installer_name": installer_name,
                "branch": installer.get("branch"),
                "total_m2": 0,
                "total_minutes": 0,
                "items_count": 0,
                "jobs": set(),
                "records": []
            }
        by_installer[installer_id]["total_m2"] += checkin.get("installed_m2", 0) or 0
        by_installer[installer_id]["total_minutes"] += duration_minutes
        by_installer[installer_id]["items_count"] += 1
        by_installer[installer_id]["jobs"].add(job.get("id"))
    
    def calc_productivity(total_m2, total_minutes):
        if total_minutes > 0 and total_m2 > 0:
            hours = total_minutes / 60
            return round(total_m2 / hours, 2)
        return 0
    
    # Prepare installer results
    installer_results = []
    for data in by_installer.values():
        data["jobs"] = list(data["jobs"])
        data["jobs_count"] = len(data["jobs"])
        data["productivity_m2_h"] = calc_productivity(data["total_m2"], data["total_minutes"])
        data["avg_minutes_per_m2"] = round(data["total_minutes"] / data["total_m2"], 2) if data["total_m2"] > 0 else 0
        data["total_hours"] = round(data["total_minutes"] / 60, 2)
        data["total_m2"] = round(data["total_m2"], 2)
        data["records"] = data["records"][:50]
        installer_results.append(data)
    
    installer_results.sort(key=lambda x: x["productivity_m2_h"], reverse=True)
    
    # Prepare job results
    job_results = []
    for data in by_job.values():
        data["installers"] = list(data["installers"])
        data["installers_count"] = len(data["installers"])
        data["productivity_m2_h"] = calc_productivity(data["total_m2_executed"], data["total_minutes"])
        data["completion_percent"] = round((data["total_m2_executed"] / data["total_m2_api"]) * 100, 1) if data["total_m2_api"] > 0 else 0
        data["total_hours"] = round(data["total_minutes"] / 60, 2)
        data["total_m2_executed"] = round(data["total_m2_executed"], 2)
        data["records"] = data["records"][:50]
        job_results.append(data)
    
    job_results.sort(key=lambda x: x["total_m2_executed"], reverse=True)
    
    # Prepare family results
    family_results = []
    for data in by_family.values():
        data["jobs"] = list(data["jobs"])
        data["installers"] = list(data["installers"])
        data["jobs_count"] = len(data["jobs"])
        data["installers_count"] = len(data["installers"])
        data["productivity_m2_h"] = calc_productivity(data["total_m2"], data["total_minutes"])
        data["avg_minutes_per_m2"] = round(data["total_minutes"] / data["total_m2"], 2) if data["total_m2"] > 0 else 0
        data["total_hours"] = round(data["total_minutes"] / 60, 2)
        data["total_m2"] = round(data["total_m2"], 2)
        data["records"] = data["records"][:50]
        family_results.append(data)
    
    family_results.sort(key=lambda x: x["total_m2"], reverse=True)
    
    # Prepare item results
    item_results = []
    for data in by_item.values():
        data["installers"] = list(data["installers"])
        data["installers_count"] = len(data["installers"])
        data["productivity_m2_h"] = calc_productivity(data["m2_api"], data["total_minutes"])
        data["avg_minutes_per_execution"] = round(data["total_minutes"] / data["executions"], 2) if data["executions"] > 0 else 0
        data["total_hours"] = round(data["total_minutes"] / 60, 2)
        data["records"] = data["records"][:20]
        item_results.append(data)
    
    item_results.sort(key=lambda x: x["m2_api"], reverse=True)
    
    total_m2 = sum(i["total_m2"] for i in installer_results)
    total_minutes = sum(i["total_minutes"] for i in by_installer.values())
    total_hours = round(total_minutes / 60, 2)
    
    return {
        "summary": {
            "total_m2": round(total_m2, 2),
            "total_hours": total_hours,
            "total_items": sum(i["items_count"] for i in installer_results),
            "total_jobs": len(by_job),
            "total_installers": len(by_installer),
            "avg_productivity_m2_h": calc_productivity(total_m2, total_minutes),
            "avg_minutes_per_m2": round(total_minutes / total_m2, 2) if total_m2 > 0 else 0,
            "filters_applied": {
                "filter_by": filter_by,
                "filter_id": filter_id,
                "date_from": date_from,
                "date_to": date_to
            }
        },
        "by_installer": installer_results if not filter_by or filter_by == "installer" else [],
        "by_job": job_results if not filter_by or filter_by == "job" else [],
        "by_family": family_results if not filter_by or filter_by == "family" else [],
        "by_item": item_results[:100] if not filter_by or filter_by == "item" else []
    }


@router.get("/metrics")
async def get_metrics(current_user: User = Depends(get_current_user)):
    """Get general metrics for dashboard."""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    total_jobs = await db.jobs.count_documents({"archived": {"$ne": True}})
    completed_jobs = await db.jobs.count_documents({"status": {"$in": ["completed", "finalizado"]}, "archived": {"$ne": True}})
    in_progress_jobs = await db.jobs.count_documents({"status": {"$in": ["in_progress", "instalando"]}, "archived": {"$ne": True}})
    pending_jobs = await db.jobs.count_documents({"status": {"$in": ["pending", "aguardando", "scheduled", "agendado"]}, "archived": {"$ne": True}})
    
    total_checkins = await db.checkins.count_documents({})
    completed_checkins = await db.checkins.count_documents({"status": "completed"})
    
    # Include item_checkins as well
    total_item_checkins = await db.item_checkins.count_documents({})
    completed_item_checkins = await db.item_checkins.count_documents({"status": "completed"})
    
    # Combine checkins for duration calculation
    completed_checkins_docs = await db.checkins.find({"status": "completed"}, {"duration_minutes": 1, "_id": 0}).to_list(1000)
    completed_item_checkins_docs = await db.item_checkins.find({"status": "completed"}, {"duration_minutes": 1, "_id": 0}).to_list(1000)
    all_completed = completed_checkins_docs + completed_item_checkins_docs
    avg_duration = sum(c.get('duration_minutes', 0) or 0 for c in all_completed) / len(all_completed) if all_completed else 0
    
    total_installers = await db.installers.count_documents({})
    
    return {
        "total_jobs": total_jobs,
        "completed_jobs": completed_jobs,
        "in_progress_jobs": in_progress_jobs,
        "pending_jobs": pending_jobs,
        "total_checkins": total_checkins + total_item_checkins,
        "completed_checkins": completed_checkins + completed_item_checkins,
        "avg_duration_minutes": round(avg_duration, 2),
        "total_installers": total_installers
    }


@router.get("/reports/export")
async def export_reports(current_user: User = Depends(get_current_user)):
    """Export consolidated report to Excel"""
    await require_role(current_user, [UserRole.ADMIN, UserRole.MANAGER])
    
    checkins = await db.item_checkins.find({}, {"_id": 0}).to_list(1000)
    jobs = await db.jobs.find({}, {"_id": 0}).to_list(1000)
    installers = await db.installers.find({}, {"_id": 0}).to_list(1000)
    
    logger.info(f"Exporting report: {len(checkins)} checkins, {len(jobs)} jobs, {len(installers)} installers")
    
    jobs_map = {job['id']: job for job in jobs}
    installers_map = {installer['id']: installer for installer in installers}
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Trabalhos"
    
    header_fill = PatternFill(start_color="FF1F5A", end_color="FF1F5A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        "Código Job", "Nome do Job", "Cliente", "Item/Produto", "Família",
        "Área Total (m²)", "M² Instalado", "Instalador",
        "GPS Check-in (Lat)", "GPS Check-in (Long)",
        "GPS Check-out (Lat)", "GPS Check-out (Long)",
        "Data Check-in", "Data Check-out", "Tempo (min)", "Status", "Filial"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    def get_product_family(product_name):
        if not product_name:
            return 'Outros'
        name = product_name.lower()
        if 'adesivo' in name:
            return 'Adesivos'
        if 'lona' in name or 'banner' in name:
            return 'Lonas/Banners'
        if 'chapa' in name or 'acm' in name or 'fachada' in name:
            return 'Chapas/Fachadas'
        if 'serviço' in name or 'instalação' in name or 'entrega' in name:
            return 'Serviços'
        if 'placa' in name or 'legenda' in name:
            return 'Placas/Legendas'
        if 'display' in name or 'totem' in name:
            return 'Displays/Totens'
        return 'Outros'
    
    row_num = 2
    for checkin in checkins:
        job = jobs_map.get(checkin.get('job_id'))
        installer = installers_map.get(checkin.get('installer_id'))
        
        if not job:
            continue
        
        product_name = checkin.get('product_name') or checkin.get('item_name') or ''
        if not product_name and job.get('holdprint_data', {}).get('products'):
            item_index = checkin.get('item_index', 0)
            products = job['holdprint_data']['products']
            if item_index < len(products):
                product_name = products[item_index].get('name', '')
        
        job_code = job.get('holdprint_data', {}).get('code') or job.get('code') or job.get('id', '')[:8]
        
        ws.cell(row=row_num, column=1, value=f"#{job_code}").border = border
        ws.cell(row=row_num, column=2, value=job.get('title', '')).border = border
        ws.cell(row=row_num, column=3, value=job.get('client_name') or job.get('holdprint_data', {}).get('customerName', '')).border = border
        ws.cell(row=row_num, column=4, value=product_name).border = border
        ws.cell(row=row_num, column=5, value=get_product_family(product_name)).border = border
        ws.cell(row=row_num, column=6, value=job.get('area_m2', 0)).border = border
        ws.cell(row=row_num, column=7, value=checkin.get('installed_m2', 0)).border = border
        ws.cell(row=row_num, column=8, value=installer.get('full_name', '') if installer else '').border = border
        ws.cell(row=row_num, column=9, value=checkin.get('gps_lat', '')).border = border
        ws.cell(row=row_num, column=10, value=checkin.get('gps_long', '')).border = border
        ws.cell(row=row_num, column=11, value=checkin.get('checkout_gps_lat', '')).border = border
        ws.cell(row=row_num, column=12, value=checkin.get('checkout_gps_long', '')).border = border
        
        checkin_at = checkin.get('checkin_at')
        if isinstance(checkin_at, str):
            try:
                checkin_at = datetime.fromisoformat(checkin_at.replace('Z', '+00:00'))
            except:
                checkin_at = None
        ws.cell(row=row_num, column=13, value=checkin_at.strftime('%d/%m/%Y %H:%M') if checkin_at else '').border = border
        
        checkout_at = checkin.get('checkout_at')
        if isinstance(checkout_at, str):
            try:
                checkout_at = datetime.fromisoformat(checkout_at.replace('Z', '+00:00'))
            except:
                checkout_at = None
        ws.cell(row=row_num, column=14, value=checkout_at.strftime('%d/%m/%Y %H:%M') if checkout_at else '').border = border
        
        ws.cell(row=row_num, column=15, value=checkin.get('duration_minutes', 0)).border = border
        ws.cell(row=row_num, column=16, value=checkin.get('status', '')).border = border
        ws.cell(row=row_num, column=17, value=job.get('branch', '')).border = border
        
        row_num += 1
    
    logger.info(f"Excel report generated with {row_num - 2} rows")
    
    column_widths = {
        'A': 12, 'B': 35, 'C': 25, 'D': 35, 'E': 18,
        'F': 15, 'G': 15, 'H': 20, 'I': 15, 'J': 15,
        'K': 15, 'L': 15, 'M': 18, 'N': 18, 'O': 12,
        'P': 15, 'Q': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"relatorio_trabalhos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
